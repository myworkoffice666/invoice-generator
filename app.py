import streamlit as st
import pandas as pd
import openpyxl
import os
import tempfile
import zipfile
from pathlib import Path

# ==========================================
# 1. 頁面配置
# ==========================================
st.set_page_config(page_title="顧問發票自動生成系統", page_icon="⚙️", layout="wide")

st.title("⚙️ 顧問發票自動生成系統")
st.markdown("""
本系統協助您將三個不同月份的顧問資料合併，並生成 Excel 發票。
請在左側欄位 **手動輸入** 本次要處理的月份標籤與評估期間。
""")

# ==========================================
# 2. 側邊欄：參數設定
# ==========================================
st.sidebar.header("📝 1. 參數設定 (請手動填寫)")

# --- 1. 評估期間輸入 ---
default_eval = '07/01/2025 - 09/30/2025'
EVALUATION_PERIOD = st.sidebar.text_input("評估期間 (Evaluation Period)", value=default_eval)

st.sidebar.markdown("---")

# --- 2. 月份標籤輸入 (P1, P2, P3) ---
st.sidebar.subheader("月份標籤設定")
st.sidebar.caption("請輸入 Excel 對應的月份名稱，這會顯示在發票上。")
P1_LABEL = st.sidebar.text_input("第 1 個月標籤 (P1)", value='Jul 2025')
P2_LABEL = st.sidebar.text_input("第 2 個月標籤 (P2)", value='Aug 2025')
P3_LABEL = st.sidebar.text_input("第 3 個月標籤 (P3)", value='Sep 2025')

st.sidebar.markdown("---")

# --- 3. 檔案上傳 (動態標籤) ---
st.sidebar.header("📂 2. 上傳檔案")
st.sidebar.caption("請依照上述設定的月份順序上傳對應檔案。")

uploaded_file_1 = st.sidebar.file_uploader(f"上傳檔案 1 ({P1_LABEL})", type=["xls", "xlsx"])
uploaded_file_2 = st.sidebar.file_uploader(f"上傳檔案 2 ({P2_LABEL})", type=["xls", "xlsx"])
uploaded_file_3 = st.sidebar.file_uploader(f"上傳檔案 3 ({P3_LABEL})", type=["xls", "xlsx"])

st.sidebar.markdown("---")
st.sidebar.header("📄 3. 上傳模板")
uploaded_template = st.sidebar.file_uploader("上傳發票模板 (CF_template.xlsx)", type=["xlsx"])

# ==========================================
# 3. 核心邏輯
# ==========================================

# 模板映射定義 (維持不變)
DATA_TEMPLATE_MAPPING = [
    (1, "D12:E12", True),  (2, "D14:E14", True),  (3, "A7:F7", True),    
    (4, "D11:E11", True),  (5, "D13:E13", True),  (6, "A5:F5", True),    
    (7, "A8:F8", True),    (8, "A16:F16", True),
    (9, "B18:B18", False), (10, "C18:C18", False), (11, "D18:D18", False), (12, "E18:E18", False),
    (13, "B19:B19", False), (14, "C19:C19", False), (15, "D19:D19", False), (16, "E19:E19", False),
    (17, "B20:B20", False), (18, "C20:C20", False), (19, "D20:D20", False), (20, "E20:E20", False),
    (21, "E21:E21", False)
]

def process_data_streamlit(files_config):
    """
    讀取並處理資料
    """
    dfs = []
    
    # 定義內部讀取函數
    def load_and_clean(file_obj, date_label):
        try:
            # Pandas 可以直接讀取 UploadedFile 物件
            df = pd.read_excel(file_obj, index_col=1, skiprows=6).iloc[:, 1:]
            
            # [優化] 移除欄位名稱的空白 (防止 'Fee ' 對應不到 'Fee')
            df.columns = df.columns.str.strip()
            
            # 安全檢查：確保有 Advisor 欄位
            if "Advisor" in df.columns:
                df = df.loc[~df["Advisor"].isna()]      # 去除空值
                df = df.loc[df["Advisor"] != "Advisor"] # 去除重複標題
            
            df["Date"] = date_label
            return df
        except Exception as e:
            st.error(f"讀取錯誤 ({date_label}): {e}")
            return pd.DataFrame()

    # 依序讀取
    for item in files_config:
        if item['file'] is not None:
            dfs.append(load_and_clean(item['file'], item['label']))
    
    if not dfs:
        return pd.DataFrame()

    # 合併
    all_data = pd.concat(dfs, axis=0, ignore_index=False).reset_index()
    if 'index' in all_data.columns:
        all_data.rename(columns={'index': 'Client'}, inplace=True)
    
    # 再次確保所有欄位去空白
    all_data.columns = all_data.columns.str.strip()

    # --- [重要] 資料清洗：處理 Excel 中的 '$' 和 ',' ---
    cols_to_clean = ['Fee', 'Average Daily Balance']
    for col in cols_to_clean:
        if col in all_data.columns:
            # 先轉字串，取代掉符號，再轉回數字
            all_data[col] = all_data[col].astype(str).str.replace(r'[$,]', '', regex=True)
            all_data[col] = pd.to_numeric(all_data[col], errors='coerce').fillna(0)
    # -----------------------------------------------------

    target_col = 'Client'
    if target_col not in all_data.columns:
        st.error(f"錯誤：在檔案中找不到 '{target_col}' 欄位，請檢查 Excel 格式。")
        return pd.DataFrame()

    # 計算出現次數，只保留剛好出現 3 次的 (因為有 P1, P2, P3)
    all_data['count'] = all_data.groupby(target_col)[target_col].transform('count')
    df_exact_3 = all_data[all_data['count'] == 3].copy()
    
    # =========================================================
    # [修改部分]：找出被排除的人，並顯示詳細缺失月份
    # =========================================================
    df_others = all_data[all_data['count'] != 3].copy()
    
    if not df_others.empty:
        # 1. 取得該次所有應該要有的月份 (Expected Months)
        expected_months_set = {item['label'] for item in files_config if item['label']}
        
        unique_excluded = df_others['Client'].unique()
        st.warning(f"⚠️ 發現 {len(unique_excluded)} 位客戶資料不完整 (非 3 個月)，已自動排除。")
        
        # 2. 整理詳細清單
        missing_details = []
        for client_name in unique_excluded:
            # 找出這個客戶目前有的資料
            client_rows = df_others[df_others['Client'] == client_name]
            present_months = set(client_rows['Date'].unique())
            
            # 找出缺少的月份 (集合相減)
            missing_months = expected_months_set - present_months
            
            missing_details.append({
                "Client (客戶名稱)": client_name,
                "Missing (缺失月份檔案)": ", ".join(missing_months) if missing_months else "Unknown",
                "Found (現有月份)": ", ".join(present_months)
            })
        
        # 3. 顯示成表格
        df_missing_report = pd.DataFrame(missing_details)
        with st.expander("📋 點擊展開：查看缺失資料詳情 (排除名單)"):
            st.dataframe(df_missing_report, use_container_width=True)
    # =========================================================

    if df_exact_3.empty:
        st.error("❌ 沒有發現剛好 3 筆資料的客戶，無法進行合併。")
        return pd.DataFrame()

    # --- Pivot 轉換 (轉寬表格) ---
    # 建立期數編號 (1, 2, 3) - 這會依據我們 append 到 dfs 的順序
    df_exact_3['period_id'] = df_exact_3.groupby(target_col).cumcount() + 1
    
    fixed_cols = ['Client', 'Advisor', 'Unique Client ID']
    # 確保這些欄位存在 (防呆)
    fixed_cols = [c for c in fixed_cols if c in df_exact_3.columns]
    
    value_cols = ['Average Daily Balance', 'Days in Period', 'Fee', 'Date']
    
    # 執行樞紐分析
    df_wide = df_exact_3.pivot(index=fixed_cols, columns='period_id', values=value_cols)
    
    # [重要步驟] 扁平化欄位名稱
    # 例如: ('Average Daily Balance', 1) -> 'Average Daily Balance1'
    # 這邊我們保留原始字串，不加底線，確保後續 generate_invoices 可以用原名抓取
    df_wide.columns = [f'{col[0]}{col[1]}' for col in df_wide.columns]
    df_wide = df_wide.reset_index()

    # 欄位整理 (保留需要的欄位)
    desired_columns = [
        'Client', 'Advisor', 'Unique Client ID',
        'Average Daily Balance1', 'Average Daily Balance2', 'Average Daily Balance3',
        'Days in Period1', 'Days in Period2', 'Days in Period3',
        'Fee1', 'Fee2', 'Fee3',
        'Date1', 'Date2', 'Date3'
    ]
    # 這裡做交集，防止找不到欄位報錯
    final_cols = [c for c in desired_columns if c in df_wide.columns]
    df_wide = df_wide[final_cols]
    
    # --- 終極防呆：計算前再次確保 Fee 是數字 ---
    for fee_col in ["Fee1", "Fee2", "Fee3"]:
        if fee_col in df_wide.columns:
            df_wide[fee_col] = pd.to_numeric(df_wide[fee_col], errors='coerce').fillna(0)

    # 計算總和
    df_wide["Total"] = (df_wide.get("Fee1", 0) + df_wide.get("Fee2", 0) + df_wide.get("Fee3", 0)).round(2)

    # 使用使用者輸入的 EVALUATION_PERIOD
    df_wide["Eval"] = EVALUATION_PERIOD

    return df_wide

def generate_invoices_streamlit(df, template_path, output_dir):
    """生成 Excel 發票"""
    xlsx_dir = Path(output_dir) / "XLSX"
    xlsx_dir.mkdir(parents=True, exist_ok=True)
    
    generated_files = []
    progress_bar = st.progress(0)
    total_rows = len(df)
    
    # [關鍵修正]：改用 to_dict('records') 而不是 itertuples
    records = df.to_dict('records')
    
    for idx, row in enumerate(records):
        # 使用 .get() 安全獲取欄位，並提供預設值
        Client = row.get("Client", "Unknown")
        
        # 處理 Unique ID (有時候會有型別問題)
        raw_id = row.get("Unique Client ID", "")
        Unique_Client_ID = str(raw_id) if pd.notna(raw_id) else ""
        
        avg1 = row.get("Average Daily Balance1", 0)
        avg2 = row.get("Average Daily Balance2", 0)
        avg3 = row.get("Average Daily Balance3", 0)
        
        days1 = row.get("Days in Period1", 0)
        days2 = row.get("Days in Period2", 0)
        days3 = row.get("Days in Period3", 0)
        
        fee1 = row.get("Fee1", 0)
        fee2 = row.get("Fee2", 0)
        fee3 = row.get("Fee3", 0)
        
        date1 = row.get("Date1", "")
        date2 = row.get("Date2", "")
        date3 = row.get("Date3", "")
        
        Total = row.get("Total", 0)
        Eval = row.get("Eval", "")

        # 準備寫入模板的資料 (21 個欄位)
        template_data = [
            # 1-8 Header
            Eval,                                       # 1
            f"${Total:,.2f}",                           # 2
            f"Client Name(s): {Client}",                # 3
            str(Unique_Client_ID)[:10],                 # 4
            "0.25%",                                    # 5
            f"Billing Cycle: {Eval}",                   # 6
            "Address: ????",                            # 7
            f"Fee Calculation {str(Unique_Client_ID)[:10]}", # 8
            
            # 9-20 Content Rows
            date1, avg1, days1, f"${fee1:,.2f}",    # Row 18
            date2, avg2, days2, f"${fee2:,.2f}",    # Row 19
            date3, avg3, days3, f"${fee3:,.2f}",    # Row 20
            
            # 21 Footer
            f"${Total:,.2f}"                            # 21
        ]

        # 處理檔名中的特殊字元
        safe_client_name = str(Client).replace("/", "_").replace("\\", "_")
        output_path = xlsx_dir / f"CF_invoice_{safe_client_name}.xlsx"
        
        try:
            # 讀取模板並填入
            wb = openpyxl.load_workbook(template_path)
            ws = wb.active
            
            for i, mapping in enumerate(DATA_TEMPLATE_MAPPING):
                index, cell_range, is_merged = mapping
                val = template_data[i]
                
                top_left = cell_range.split(':')[0]
                if is_merged:
                    try: ws.merge_cells(cell_range)
                    except ValueError: pass
                ws[top_left] = val
            
            wb.save(output_path)
            generated_files.append(output_path)
        except Exception as e:
            st.error(f"生成失敗 ({Client}): {e}")
        
        progress_bar.progress((idx + 1) / total_rows)
        
    return generated_files

def make_zip(source_dirs, output_filename):
    """將資料夾打包成 ZIP"""
    zip_path = Path(output_filename)
    with zipfile.ZipFile(zip_path, 'w', zipfile.ZIP_DEFLATED) as zipf:
        for folder in source_dirs:
            folder_path = Path(folder)
            if folder_path.exists():
                for file in folder_path.glob('*'):
                    zipf.write(file, arcname=f"{folder_path.name}/{file.name}")
    return zip_path

# ==========================================
# 4. 主執行流程
# ==========================================

# 顯示當前設定摘要
st.markdown(f"""
### 📋 當前處理設定
| 參數 | 設定值 |
| :--- | :--- |
| **評估期間** | `{EVALUATION_PERIOD}` |
| **Period 1** | `{P1_LABEL}` |
| **Period 2** | `{P2_LABEL}` |
| **Period 3** | `{P3_LABEL}` |
""")

start_button = st.sidebar.button("🚀 開始處理", type="primary")

if start_button:
    # 檢查檔案是否齊全
    if not (uploaded_file_1 and uploaded_file_2 and uploaded_file_3 and uploaded_template):
        st.error("請先上傳所有必要的檔案 (3個月份資料 + 1個模板)。")
    else:
        # 建立臨時工作目錄
        with tempfile.TemporaryDirectory() as tmpdirname:
            st.info(f"環境準備就緒，開始運算...")
            
            # 1. 儲存模板到臨時目錄
            temp_template_path = os.path.join(tmpdirname, "template.xlsx")
            with open(temp_template_path, "wb") as f:
                f.write(uploaded_template.getbuffer())
            
            # 2. 準備檔案列表 (包含動態標籤)
            files_config = [
                {'file': uploaded_file_1, 'label': P1_LABEL},
                {'file': uploaded_file_2, 'label': P2_LABEL},
                {'file': uploaded_file_3, 'label': P3_LABEL}
            ]
            
            # Step 1: 資料處理
            with st.spinner('Step 1: 正在讀取並合併資料...'):
                df_result = process_data_streamlit(files_config)
            
            if not df_result.empty:
                st.success(f"資料處理完成！共 {len(df_result)} 位合格客戶。")
                
                # 預覽數據
                st.subheader("📊 數據預覽 (請確認 Fee, Balance, Date 是否正確)")
                preview_cols = ['Client', 'Total', 'Fee1', 'Fee2', 'Fee3', 'Date1', 'Date2', 'Date3', 'Average Daily Balance1']
                # 只顯示存在的欄位
                valid_preview = [c for c in preview_cols if c in df_result.columns]
                st.dataframe(df_result[valid_preview].head(10))
                
                # Step 2: 生成 Excel
                xlsx_output_dir = os.path.join(tmpdirname, "XLSX")
                with st.spinner('Step 2: 正在生成 Excel 發票...'):
                    generated_xlsx = generate_invoices_streamlit(df_result, temp_template_path, tmpdirname)
                
                st.success(f"已生成 {len(generated_xlsx)} 份 Excel 發票。")
                
                # Step 3: 打包下載
                with st.spinner('正在打包檔案...'):
                    dirs_to_zip = [xlsx_output_dir]
                    zip_filename = os.path.join(tmpdirname, "invoices_result.zip")
                    zip_path = make_zip(dirs_to_zip, zip_filename)
                    
                    with open(zip_path, "rb") as f:
                        zip_data = f.read()
                        
                    st.balloons()
                    st.header("🎉 處理完成！")
                    st.download_button(
                        label="📥 下載完整壓縮包 (ZIP)",
                        data=zip_data,
                        file_name=f"invoices_{P1_LABEL}_to_{P3_LABEL}.zip",
                        mime="application/zip"
                    )
            else:
                st.warning("沒有產生任何數據，請檢查上傳的檔案內容。")

st.markdown("---")
st.caption("Powered by Streamlit & Python")
